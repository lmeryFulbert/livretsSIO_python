import openpyxl     #faire un pip install openpyxl pour installer la librairie
import datetime
import json

#constante pour le dossier correspondant à l'année de travail
annee = datetime.datetime.now().year


# Spécifier le chemin relatif du fichier etudiants
file_in_pronote_etudiants = 'in/'+ str(annee) +'/etudiants.xlsx'

# Spécifier le chemin relatif du fichier de resultats de l'année 1
file_in_pronote_an1 = 'in/'+ str(annee) +'/an1.xlsx'
# Spécifier le chemin relatif des resultat de l'an2 et des commentaires à placer sur le livret
file_in_pronote_an2 = 'in/'+ str(annee) +'/an2.xlsx'

# Spécifier le chemin relatif du fichier des disciplines
file_in_disciplines = 'in/'+ str(annee) +'/discipline.xlsx'

# ouverture du fichier contenant les étudiants
workbook_livret_pronote_etudiants = openpyxl.load_workbook(file_in_pronote_etudiants,data_only=True)

# ouverture du fichier contenant les disciplines
workbook_discipline = openpyxl.load_workbook(file_in_disciplines)

# ouverture des contenant les notes des étudiants
workbook_livret_pronote_an1 = openpyxl.load_workbook(file_in_pronote_an1,data_only=True)
workbook_livret_pronote_an2 = openpyxl.load_workbook(file_in_pronote_an2)

# Accéder à la feuille des étudiants
etuiants = workbook_livret_pronote_etudiants['Feuil1']

# ouvertur fichier résultats an1
sem1=workbook_livret_pronote_an1['SEMESTRE1']
sem2=workbook_livret_pronote_an1['SEMESTRE2']
an1=workbook_livret_pronote_an1['SIO1ANNEE']

#recuperation des noms de matieres pour les transformer en nom de bloc
feuille_discipline=workbook_discipline['SIO1ANNEE']
feuille2_discipline=workbook_discipline['SIO2ANNEE']

disciplines=dict()

for row in feuille_discipline.iter_rows(min_row=1, values_only=True):
    disciplines[row[0]]=row[1].strip()

for row in feuille2_discipline.iter_rows(min_row=1, values_only=True):
    disciplines[row[0]]=row[1].strip()

#print(disciplines)

data = []

i=1
# Parcourir la liste des étudiants
for row in etuiants.iter_rows(min_row=2, values_only=True):
    #print("etudiant " + str(i))
    row_data = {
        'nom': row[0],
        'prenom':row[1],
        'date_naiss': row[2].strftime("%d-%m-%Y"),
        'pix': row[6],
        'specialite': row[7],
        'avis': row[8]
    }
    #
    row_data['an1'] = []     # Créer une liste vide pour les lignes de l'année 1

    old_discipline=''

    #on cherche le code des matières a partir de la colonne 4
    #parcours des moyennes annuelles
    for colonne in an1.iter_cols(min_col=4, values_only=True):
        #print(colonne[0])
        matiere=colonne[0]
        #recherche de la moyenne su semestre 1
        for colsem1 in sem1.iter_cols(min_col=4, values_only=True):
            if colsem1[0]==matiere:
                #recuperation de la moyenne du semestre 1
                #print(str(len(colsem1))  + "  idetud : " + str(i))
                #print(colsem1)
                note_sem1=colsem1[i+1]
                break

        #recherche de la moyenne su semestre 2
        for colsem2 in sem2.iter_cols(min_col=4, values_only=True):
            #print(colsem2[0])
            if colsem2[0]==matiere:
                #recuperation de la moyenne du semestre 2
                note_sem2=colsem2[i+1]
                break

        # si le contenu de la cellule est bien un nombre
        if colonne[i+1] is not None and (isinstance(colonne[i+1], float) or isinstance(colonne[i+1], int)): 
            #recuperation de la moyenne annuelle
            note_an1=colonne[i+1]


       # print("etudiant " + str(i) +" : " + str(disciplines.get(colonne[0])))
                 
        if(old_discipline!=disciplines.get(colonne[0])):
            old_discipline=disciplines.get(colonne[0])

            feuillecalcul=workbook_livret_pronote_an1['CALCUL']

            if(disciplines.get(colonne[0])=='B1'):
                note_sem1=feuillecalcul.cell(i+2, 5).value
                note_sem2=feuillecalcul.cell(row=i+2, column=6).value
                if(note_sem1 is not None and note_sem2 is not None):
                    note_an1=(note_sem1 + note_sem2)/2
                else:
                    note_an1=None

            if(disciplines.get(colonne[0])=='B3'):
                note_sem1=feuillecalcul.cell(row=i+2, column=7).value
                note_sem2=feuillecalcul.cell(row=i+2, column=8).value
                if(note_sem1 is not None and note_sem2 is not None):
                    note_an1=(note_sem1 + note_sem2)/2
                else:
                    note_an1=None

            if(disciplines.get(colonne[0])=='B2'):
                note_sem1= None
                note_sem2=feuillecalcul.cell(row=i+2, column=13).value
                note_an1=note_sem2

            col_an1 = {
                'discipline': disciplines.get(colonne[0]),
                'moyenne_sem1': note_sem1,
                'moyenne_sem2': note_sem2,
                'moyenne_an1': note_an1
            }

            #print(col_an1)
            note_an1=None

            # Ajouter la ligne du livret à la liste des lignes du livret
            row_data['an1'].append(col_an1) 
    
    # resultat AN2 + commentaire
    # Accéder à la feuille du livret de l'étudiant
    sheetname='Feuil' + str(i)
    livret_scolaire = workbook_livret_pronote_an2[sheetname]

    row_data['an2'] = []     # Créer une liste vide pour les lignes dd l'année 2 et les copmmentaires
    for ligne in livret_scolaire.iter_rows(min_row=2, values_only=True):

        if(ligne[1]=='Sem1'):
            note_sem1=ligne[3]

        if(ligne[1]=='Sem2'):
            note_sem2=ligne[3]

        if(ligne[1]=='Année'):
            note_an2=ligne[3]

            moy_classe=ligne[4]

            row_an2 = {
                'discipline': disciplines.get(ligne[0]),
                'moy_sem1': note_sem1,
                'moy_sem2': note_sem2,
                'moy_an2': note_an2,
                'moy_classe': moy_classe,
                'commentaire':ligne[8]
            }
            # Ajouter la ligne du livret à la liste des lignes du livret
            row_data['an2'].append(row_an2) 
            note_an2=None

    i=i+1
    # Ajouter le dictionnaire à la liste des données
    data.append(row_data)

# Créer un dictionnaire pour les statistiques

statistics = {}
statistics['TF'] = str(round(etuiants['M2'].value * 100)) + '%'
statistics['F'] = str(round(etuiants['M3'].value * 100)) + '%'
statistics['P'] = str(round(etuiants['M4'].value * 100)) + '%'

listelivrets = {
                'statistiques': statistics,
                'livrets': data
                }
    
# # Convertir le dictionnaire en objet JSON
json_data = json.dumps(listelivrets, indent=4, ensure_ascii=False).encode('utf8')

# # Enregistrer le fichier JSON
with open('json/data.json', 'wb') as f:
    f.write(json_data)

